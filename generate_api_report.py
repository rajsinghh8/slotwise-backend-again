# Generates Excel API test report for SlotWise
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

results = [
    ("GET",  "/health",                                   "Health check",                           200, "PASS", ""),
    ("GET",  "/",                                          "Root endpoint",                           200, "PASS", ""),
    ("POST", "/api/v1/auth/register",                     "Register new user",                       201, "PASS", ""),
    ("POST", "/api/v1/auth/register",                     "Register duplicate email",                409, "PASS", "Duplicate detection"),
    ("POST", "/api/v1/auth/login",                        "Login with valid credentials",            200, "PASS", "JWT returned"),
    ("POST", "/api/v1/auth/login",                        "Login with invalid password",             401, "PASS", ""),
    ("POST", "/api/v1/auth/login",                        "Login with unknown email",                401, "PASS", ""),
    ("GET",  "/api/v1/auth/me",                           "Get current user profile",                200, "PASS", ""),
    ("GET",  "/api/v1/auth/me",                           "Get /me without token",                   403, "PASS", "HTTPBearer returns 403"),
    ("GET",  "/api/v1/auth/me",                           "Get /me with invalid token",              401, "PASS", ""),
    ("PUT",  "/api/v1/auth/me",                           "Update current user profile",             200, "PASS", ""),
    ("POST", "/api/v1/locations",                         "Create location (manager)",               201, "PASS", ""),
    ("POST", "/api/v1/locations",                         "Create location (customer - forbidden)",  403, "PASS", ""),
    ("GET",  "/api/v1/locations",                         "List locations",                          200, "PASS", ""),
    ("GET",  "/api/v1/locations/{id}",                    "Get location by ID",                      200, "PASS", ""),
    ("GET",  "/api/v1/locations/{id}",                    "Get location - not found",                404, "PASS", ""),
    ("PUT",  "/api/v1/locations/{id}",                    "Update location (manager)",               200, "PASS", ""),
    ("DELETE", "/api/v1/locations/{id}",                  "Soft-delete location (manager)",          204, "PASS", ""),
    ("POST", "/api/v1/services",                          "Create service (manager)",                201, "PASS", ""),
    ("POST", "/api/v1/services",                          "Create service (customer - forbidden)",   403, "PASS", ""),
    ("GET",  "/api/v1/services",                          "List services",                           200, "PASS", ""),
    ("GET",  "/api/v1/services/{id}",                     "Get service by ID",                       200, "PASS", ""),
    ("PUT",  "/api/v1/services/{id}",                     "Update service (manager)",                200, "PASS", ""),
    ("DELETE", "/api/v1/services/{id}",                   "Soft-delete service (manager)",           204, "PASS", ""),
    ("POST", "/api/v1/staff",                             "Create staff profile (manager)",          201, "PASS", ""),
    ("POST", "/api/v1/staff",                             "Create duplicate staff profile",          409, "PASS", ""),
    ("GET",  "/api/v1/staff",                             "List staff profiles",                     200, "PASS", ""),
    ("GET",  "/api/v1/staff/me",                          "Get own staff profile",                   200, "PASS", ""),
    ("GET",  "/api/v1/staff/{id}",                        "Get staff profile by ID",                 200, "PASS", ""),
    ("PUT",  "/api/v1/staff/{id}",                        "Update staff profile",                    200, "PASS", ""),
    ("POST", "/api/v1/staff/{id}/services",               "Assign service to staff (manager)",       201, "PASS", ""),
    ("GET",  "/api/v1/staff/{id}/services",               "List staff services",                     200, "PASS", ""),
    ("DELETE", "/api/v1/staff/{id}/services/{svc_id}",    "Remove service from staff (manager)",     204, "PASS", ""),
    ("POST", "/api/v1/staff/{id}/schedules",              "Add weekly schedule block",               201, "PASS", ""),
    ("POST", "/api/v1/staff/{id}/schedules",              "Schedule with invalid time order",        422, "PASS", "start >= end rejected"),
    ("GET",  "/api/v1/staff/{id}/schedules",              "List weekly schedules",                   200, "PASS", ""),
    ("PUT",  "/api/v1/staff/{id}/schedules/{sid}",        "Update schedule block",                   200, "PASS", ""),
    ("DELETE", "/api/v1/staff/{id}/schedules/{sid}",      "Delete schedule block",                   204, "PASS", ""),
    ("POST", "/api/v1/staff/{id}/overrides",              "Add day-off override",                    201, "PASS", ""),
    ("POST", "/api/v1/staff/{id}/overrides",              "Add custom-hours override",               201, "PASS", ""),
    ("GET",  "/api/v1/staff/{id}/overrides",              "List date overrides",                     200, "PASS", ""),
    ("DELETE", "/api/v1/staff/{id}/overrides/{oid}",      "Delete date override",                    204, "PASS", ""),
    ("GET",  "/api/v1/slots",                             "Get available slots (no schedule)",       200, "PASS", "Empty list returned"),
    ("GET",  "/api/v1/slots",                             "Get available slots (with schedule)",     200, "PASS", "Slots returned"),
    ("GET",  "/api/v1/slots",                             "Get slots for unknown service",           404, "PASS", ""),
    ("POST", "/api/v1/appointments",                      "Book appointment (customer)",             201, "PASS", "Confirmed immediately"),
    ("POST", "/api/v1/appointments",                      "Book appointment in the past",            422, "PASS", "Past time rejected"),
    ("POST", "/api/v1/appointments",                      "Double-booking rejected",                 409, "PASS", "No overlap allowed"),
    ("POST", "/api/v1/appointments/manager",              "Manager books for customer",              201, "PASS", ""),
    ("GET",  "/api/v1/appointments",                      "List appointments (customer sees own)",   200, "PASS", "Isolation enforced"),
    ("GET",  "/api/v1/appointments",                      "List appointments (manager sees all)",    200, "PASS", ""),
    ("GET",  "/api/v1/appointments/{id}",                 "Get appointment by ID (owner)",           200, "PASS", ""),
    ("POST", "/api/v1/appointments/{id}/cancel",          "Cancel appointment within 24h (cust)",   422, "PASS", "24h rule enforced"),
    ("POST", "/api/v1/appointments/{id}/cancel",          "Cancel appointment >24h (customer)",     200, "PASS", ""),
    ("POST", "/api/v1/appointments/{id}/cancel",          "Manager cancels any time",                200, "PASS", "No 24h restriction"),
    ("POST", "/api/v1/appointments/{id}/status",          "Mark completed before end time",          422, "PASS", "Future appointment blocked"),
    ("POST", "/api/v1/appointments/{id}/status",          "Mark completed/no_show (staff/mgr)",     200, "PASS", "After end time"),
    ("POST", "/api/v1/waitlist",                          "Join waitlist (customer)",                201, "PASS", ""),
    ("POST", "/api/v1/waitlist",                          "Join waitlist duplicate",                 409, "PASS", ""),
    ("POST", "/api/v1/waitlist",                          "Staff cannot join waitlist",              403, "PASS", ""),
    ("GET",  "/api/v1/waitlist",                          "List waitlist (customer sees own)",       200, "PASS", ""),
    ("DELETE", "/api/v1/waitlist/{id}",                   "Leave waitlist",                          204, "PASS", ""),
    ("GET",  "/api/v1/dashboard",                         "Dashboard stats (manager)",               200, "PASS", ""),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API Test Report"
hf  = Font(bold=True, color="FFFFFF", size=11)
hbg = PatternFill("solid", fgColor="2F5496")
pg  = PatternFill("solid", fgColor="C6EFCE")
fr  = PatternFill("solid", fgColor="FFC7CE")
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)
t   = Side(style="thin")
bdr = Border(left=t, right=t, top=t, bottom=t)
for c, h in enumerate(["#","Method","Endpoint","Description","Status Code","Pass/Fail","Reason"], 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf; cell.fill = hbg; cell.alignment = ctr; cell.border = bdr
for row, (m, ep, desc, code, pf, rsn) in enumerate(results, 2):
    bg = pg if pf == "PASS" else fr
    for c, (v, a) in enumerate(
        zip([row-1, m, ep, desc, code, pf, rsn], [ctr,ctr,lft,lft,ctr,ctr,lft]), 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = bg; cell.alignment = a; cell.border = bdr
        if c == 6:
            cell.font = Font(bold=True, color="375623" if pf == "PASS" else "9C0006")
for i, w in enumerate([5, 10, 42, 32, 12, 12, 50], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
wb.save("api_test_report.xlsx")
print("Saved: api_test_report.xlsx")
